import openpyxl

def extract_excel_headers(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

def extract_excel_data(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers))}
        data_rows.append(row_data)
    return data_rows
